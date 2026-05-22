"""
file_generator.py — Build and save the PICQD output Excel + ZIP.

Output structure (mirrors SIMBA pipeline convention):
  output/<timestamp>/
      PICQD_DATA_<YYYYMMDD>.xlsx
      PICQD_DATA_<YYYYMMDD>.zip
  output/latest/
      PICQD_DATA_<YYYYMMDD>.xlsx
      PICQD_DATA_<YYYYMMDD>.zip

Excel layout:
  Row 1  : blank | COLUMN_CODE_1 | ... | COLUMN_CODE_228
  Row 2  : blank | header_1      | ... | header_228
  Row 3+ : period | value_1      | ... | value_228   (one row per period)

Values are floats where present, empty cell where None.
"""
import os
import shutil
import zipfile
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import config

logger = logging.getLogger(__name__)


# ── Internal Excel builder ────────────────────────────────────────────────────

def _write_excel(records, xlsx_path):
    """Write the Data sheet to xlsx_path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    code_font   = Font(bold=True, size=9)
    header_font = Font(italic=True, size=9)
    code_fill   = PatternFill("solid", fgColor="D9E1F2")
    header_fill = PatternFill("solid", fgColor="EBF1DE")
    centre      = Alignment(horizontal="center", vertical="top", wrap_text=True)
    top_left    = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    # Row 1 — column codes
    ws.cell(1, 1, "")
    for col_idx, code in enumerate(config.COLUMN_CODES, start=2):
        cell = ws.cell(1, col_idx, code)
        cell.font      = code_font
        cell.fill      = code_fill
        cell.alignment = centre

    # Row 2 — column headers
    ws.cell(2, 1, "")
    for col_idx, hdr in enumerate(config.COLUMN_HEADERS, start=2):
        cell = ws.cell(2, col_idx, hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = top_left

    # Data rows
    for row_idx, (period, data) in enumerate(records, start=3):
        ws.cell(row_idx, 1, period)
        for col_idx, code in enumerate(config.COLUMN_CODES, start=2):
            val = data.get(code)
            if val is not None:
                ws.cell(row_idx, col_idx, val)

    # Column widths
    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(config.COLUMN_CODES) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    ws.freeze_panes = "B3"
    wb.save(str(xlsx_path))
    logger.info("Excel written: %s", xlsx_path)


# ── Public entry point ────────────────────────────────────────────────────────

def generate(records, date_str):
    """
    Build Excel + ZIP, save to timestamped run folder and output/latest/.

    Parameters
    ----------
    records  : list of (period: str, data: dict[code -> float|None])
               sorted ascending by period before calling.
    date_str : YYYYMMDD string from the release date (e.g. '20260515').

    Returns
    -------
    run_dir  : Path to the timestamped output folder.
    """
    ts      = datetime.now().strftime('%Y%m%d_%H%M%S')
    run_dir = Path(config.OUTPUT_DIR) / ts
    run_dir.mkdir(parents=True, exist_ok=True)

    latest_dir = Path(config.OUTPUT_DIR) / 'latest'
    if latest_dir.exists():
        shutil.rmtree(str(latest_dir))
    latest_dir.mkdir(parents=True)

    base_name = f"PICQD_DATA_{date_str}"
    xlsx_name = base_name + ".xlsx"
    zip_name  = base_name + ".zip"

    xlsx_path = run_dir / xlsx_name
    zip_path  = run_dir / zip_name

    # Write Excel
    _write_excel(records, xlsx_path)

    # Wrap in ZIP
    with zipfile.ZipFile(str(zip_path), 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(str(xlsx_path), xlsx_name)
    logger.info("ZIP written: %s", zip_path)

    # Copy both to latest/
    shutil.copy2(str(xlsx_path), str(latest_dir / xlsx_name))
    shutil.copy2(str(zip_path),  str(latest_dir / zip_name))
    logger.info("Copied to latest/: %s, %s", xlsx_name, zip_name)

    filled = sum(1 for _, d in records for v in d.values() if v is not None)
    logger.info(
        "Output complete — %d period(s), %d non-null values total",
        len(records), filled,
    )
    return run_dir
